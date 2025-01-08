from flask import Flask, render_template, request, jsonify
import cv2
import os
import sqlite3
import torch
import numpy as np
from facenet_pytorch import MTCNN, InceptionResnetV1
from datetime import datetime
import openpyxl  # Import openpyxl for Excel file handling

# Initialize Flask
app = Flask(__name__)

# Initialize MTCNN and FaceNet model
mtcnn = MTCNN(keep_all=False)
model = InceptionResnetV1(pretrained='vggface2').eval()

# Database and Embedding Storage
database_file = 'attendance_system.db'
embeddings_file = 'embeddings.npz'
excel_file = 'Excel1.xlsx'  # Excel file to store attendance

# Initialize global lists for embeddings and names
embeddings = []
names = []

# Create or connect to SQLite database
def create_database_connection():
    conn = sqlite3.connect(database_file)
    cursor = conn.cursor()
    cursor.execute(''' 
        CREATE TABLE IF NOT EXISTS users ( 
            id INTEGER PRIMARY KEY, 
            name TEXT, 
            image_path TEXT 
        ) 
    ''')
    conn.commit()
    return conn, cursor

def close_database_connection(conn):
    conn.commit()
    conn.close()

# Preprocess face for model input
def preprocess_face(face):
    face_resized = cv2.resize(face, (160, 160))
    face_rgb = cv2.cvtColor(face_resized, cv2.COLOR_BGR2RGB)
    face_tensor = torch.tensor(face_rgb).permute(2, 0, 1).float().div(255.0).unsqueeze(0)
    return face_tensor

# Load embeddings from file (this runs at the start of the program)
def load_embeddings():
    global embeddings, names
    if os.path.exists(embeddings_file):
        try:
            data = np.load(embeddings_file, allow_pickle=True)
            if 'embeddings' in data and 'names' in data:
                embeddings = list(data['embeddings'])
                names = list(data['names'])
                print(f"Loaded {len(embeddings)} embeddings from {embeddings_file}.")
            else:
                print(f"File {embeddings_file} is missing required keys. Reinitializing...")
                reinitialize_embeddings_file()
        except Exception as e:
            print(f"Error loading embeddings: {e}. Reinitializing...")
            reinitialize_embeddings_file()
    else:
        print(f"{embeddings_file} not found. Initializing new file...")
        reinitialize_embeddings_file()

# Reinitialize embeddings file
def reinitialize_embeddings_file():
    global embeddings, names
    embeddings = []
    names = []
    save_embeddings()
    print(f"Reinitialized {embeddings_file} with empty data.")

# Save embeddings and names to a file
def save_embeddings():
    np.savez(embeddings_file, embeddings=np.array(embeddings), names=np.array(names))
    print(f"Saved embeddings to {embeddings_file}.")

# Initialize Excel Workbook for Attendance Logging
def initialize_excel_file():
    if not os.path.exists(excel_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["User Name", "Date", "Time"])  # Header row
        workbook.save(excel_file)
    else:
        print(f"{excel_file} already exists.")

# Write attendance to Excel file
def write_to_excel(user_name):
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")
    
    # Open existing Excel file and add a new row
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    
    # Check if the entry already exists
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] == user_name and row[1] == current_date:
            print(f"Entry for {user_name} already exists for today. Skipping...")
            return  # Skip adding duplicate entry
    
    # If entry is unique, write to the Excel file
    sheet.append([user_name, current_date, current_time])
    workbook.save(excel_file)
    print(f"Logged {user_name} attendance in {excel_file}")

# Function to capture faces and generate embeddings for a user
def capture_faces_and_generate_embeddings(conn, cursor, user_name):
    global embeddings, names
    count = 0
    user_folder = f'images/user_{user_name}'
    os.makedirs(user_folder, exist_ok=True)

    # Create webcam instance
    cap = cv2.VideoCapture(0)

    print(f"Capturing images for {user_name}...")

    while count < 5:  # Capture 5 images for simplicity
        ret, frame = cap.read()
        if not ret:
            print("Failed to capture image. Exiting...")
            break

        cv2.imshow(f"Capturing for {user_name}", frame)
        key = cv2.waitKey(1)

        if key == ord('c'):  # Press 'c' to capture an image
            # Save the image
            file_name = f"{user_folder}/image_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            cv2.imwrite(file_name, frame)
            count += 1
            print(f"Captured image {count}/5 for {user_name}")

            # Detect face and generate embedding
            boxes, _ = mtcnn.detect(frame)
            if boxes is not None:
                for box in boxes:
                    x1, y1, x2, y2 = map(int, box)
                    face = frame[y1:y2, x1:x2]

                    if face.size == 0:
                        continue

                    # Preprocess the face before passing to the model
                    face_tensor = preprocess_face(face)

                    # Generate embedding
                    with torch.no_grad():
                        embedding = model(face_tensor).numpy()
                        embeddings.append(embedding)
                        names.append(user_name)

                    # Save user data to the database
                    cursor.execute("INSERT INTO users (name, image_path) VALUES (?, ?)", (user_name, file_name))
                    conn.commit()

        elif key == ord('q'):  # Press 'q' to quit
            print("Quitting...")
            break

    # Save embeddings and names to a file
    save_embeddings()

    # Release webcam and close database
    cap.release()
    cv2.destroyAllWindows()

# Function to perform face recognition
def perform_face_recognition(conn, cursor):
    global embeddings, names

    # Load saved embeddings and names
    load_embeddings()

    # Initialize webcam
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Could not access the camera.")
        return

    print("Starting face recognition...")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Error: Failed to capture image.")
            break

        # Detect faces using MTCNN
        boxes, _ = mtcnn.detect(frame)

        if boxes is not None:
            for box in boxes:
                x1, y1, x2, y2 = map(int, box)
                face = frame[y1:y2, x1:x2]

                if face.size == 0:
                    continue

                # Preprocess the face before passing to the model
                face_tensor = preprocess_face(face)

                with torch.no_grad():
                    embedding = model(face_tensor).numpy()

                # Compare with known embeddings
                name = "Unknown"
                min_distance = float("inf")
                for stored_embedding, stored_name in zip(embeddings, names):
                    distance = np.linalg.norm(embedding - stored_embedding)
                    if distance < min_distance:
                        min_distance = distance
                        name = stored_name if distance < 0.75 else "Unknown"

                # Draw box and name for each detected face
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

                # If a recognized user is detected, log the attendance
                if name != "Unknown":
                    write_to_excel(name)

        cv2.imshow("Face Recognition", frame)

        # Exit if 'Esc' key is pressed
        key = cv2.waitKey(1) & 0xFF
        if key == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

# Flask routes
@app.route('/')
def home():
    return render_template('index.html')  # Render the main page

@app.route('/start_detection', methods=['POST'])
def start_detection():
    conn, cursor = create_database_connection()
    perform_face_recognition(conn, cursor)  # Start face detection
    close_database_connection(conn)
    return jsonify({"message": "Detection started."})

@app.route('/add_user', methods=['POST'])
def add_user():
    user_name = request.form['user_name']
    conn, cursor = create_database_connection()
    capture_faces_and_generate_embeddings(conn, cursor, user_name)  # Add new user
    close_database_connection(conn)
    return jsonify({"message": f"User {user_name} added successfully."})

if __name__ == "__main__":
    load_embeddings()  # Load embeddings at startup
    initialize_excel_file()  # Initialize Excel file at the start
    app.run(debug=True, host="0.0.0.0", port=5000)  # Run Flask app
